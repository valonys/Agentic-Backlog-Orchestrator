"""
Utility functions for Excel parsing and AI processing
"""
import openpyxl
import openpyxl.utils
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
import re
import os
import json
import logging

# Optional AI imports - gracefully handle if not installed
try:
    from langchain_openai import ChatOpenAI
    from crewai import Agent, Task, Crew
    AI_AVAILABLE = True
except ImportError:
    AI_AVAILABLE = False
    ChatOpenAI = None
    Agent = None
    Task = None
    Crew = None
    Process = None

logger = logging.getLogger(__name__)

# Valid site prefixes for filename validation
VALID_SITE_PREFIXES = ['GIR', 'DAL', 'PAZ', 'CLV']

# Excel epoch: 1899-12-30 (Windows). Serial 1 = 1899-12-31.
_EXCEL_EPOCH = datetime(1899, 12, 30)


def _excel_serial_to_date_str(value: Any) -> Optional[str]:
    """
    Convert Excel date serial (int/float) to YYYY-MM-DD string.
    Returns None if value is not a valid serial in typical range.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            serial = int(round(float(value)))
            if serial < 1 or serial > 100000:
                return None
            d = _EXCEL_EPOCH + timedelta(days=serial)
            return d.strftime('%Y-%m-%d')
        except (ValueError, OverflowError):
            return None
    return None


def validate_filename_pattern(filename: str) -> tuple[bool, str]:
    """
    Validate that filename follows the required pattern: {SITE}_*.xlsx

    Args:
        filename: Name of the uploaded file

    Returns:
        Tuple of (is_valid, site_prefix or error_message)
    """
    if not filename:
        return False, "Filename is empty"

    # Extract first 3 characters (case-insensitive)
    prefix = filename[:3].upper()

    if prefix not in VALID_SITE_PREFIXES:
        return False, f"Invalid filename pattern. File must start with one of: {', '.join(VALID_SITE_PREFIXES)}. Got: '{prefix}'"

    logger.info(f"Filename validation passed: {filename} -> Site: {prefix}")
    return True, prefix


def read_database_sheet(file_path: str) -> List[Dict[str, Any]]:
    """
    Reads the 'Data Base' sheet from Excel file with preprocessing:
    - Removes rows 1-4 (metadata/title rows)
    - Removes column A (unused/index column)
    - Row 5 becomes the header row (after removal, this is the first data row)
    - Data starts from row 6 onwards
    - Due Date column is formatted as proper date

    NOTE: File must follow naming pattern: {SITE}_*.xlsx where SITE in [GIR, DAL, PAZ, CLV]

    Args:
        file_path: Path to Excel file

    Returns:
        List of dictionaries containing backlog items

    Raises:
        ValueError: If file cannot be parsed or sheet not found
    """
    try:
        # Load workbook
        logger.info(f"Loading workbook: {file_path}")
        # Use read_only=False for xlsm files to avoid hanging
        # data_only=True converts formulas to values
        wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True, keep_vba=False)
        logger.info(f"Workbook loaded successfully with preprocessing mode")
        
        # Target ONLY the "Data Base" worksheet (case-insensitive, allows "DataBase")
        target_name = None
        for name in wb.sheetnames:
            if re.fullmatch(r"data\s*base", name.strip(), re.IGNORECASE) or re.fullmatch(r"database", name.strip(), re.IGNORECASE):
                target_name = name
                break

        if not target_name:
            available = ", ".join(wb.sheetnames)
            raise ValueError(f"No sheet named 'Data Base' found. Available: {available}")

        logger.info(f"Found sheet: {target_name}")
        ws = wb[target_name]
        logger.info(f"Accessing sheet data...")
        # Dynamically detect last column with a header in row 5 (up to col 40 / column AN)
        MAX_COL_SCAN = 40
        last_col = 24  # minimum fallback (column X)
        for col in range(2, MAX_COL_SCAN + 1):
            if ws.cell(row=5, column=col).value is not None:
                last_col = col
        logger.info(f"Detected header columns B(2) to {last_col} ({openpyxl.utils.get_column_letter(last_col)})")

        header_cells = [ws.cell(row=5, column=col).value for col in range(2, last_col + 1)]
        if not any(header_cells):
            raise ValueError(f"Sheet 'Data Base' does not contain headers in B5:{openpyxl.utils.get_column_letter(last_col)}5")

        headers = [str(c).strip() if c is not None else f"Col_{idx}"
                   for idx, c in enumerate(header_cells)]
        canonical_headers = map_headers(headers)
        # Log column mapping for Backlog? and Due Date (helps debug wrong counts/dates)
        for idx, h in enumerate(canonical_headers):
            if h == 'Backlog?':
                logger.info(f"Column index {idx} (0-based from B) mapped to 'Backlog?' (raw header: {headers[idx]})")
            if h == 'Due Date':
                logger.info(f"Column index {idx} (0-based from B) mapped to 'Due Date' (raw header: {headers[idx]})")
        if 'Backlog?' not in canonical_headers:
            logger.warning("No column mapped to 'Backlog?' — backlog counts may be wrong. Check for header containing 'Backlog'.")
        if 'Due Date' not in canonical_headers:
            logger.warning("No column mapped to 'Due Date' — check for header containing 'Due Date'.")

        # Optimized: Read rows in batches using iter_rows (faster than cell-by-cell)
        logger.info(f"Starting to parse rows from row 6...")
        records: List[Dict[str, Any]] = []
        empty_streak = 0
        max_empty_streak = 50  # Reduced for faster detection
        row_idx = 6
        rows_processed = 0
        
        # Use iter_rows for faster batch reading (columns B to detected last_col)
        for row in ws.iter_rows(min_row=6, min_col=2, max_col=last_col, values_only=True):
            # Check if row has ANY non-empty cells
            has_data = any(v is not None and str(v).strip() != '' for v in row)

            if not has_data:
                empty_streak += 1
                if empty_streak >= max_empty_streak:
                    logger.info(f"Stopping at row {row_idx} after {empty_streak} consecutive empty rows")
                    break
                row_idx += 1
                continue

            empty_streak = 0

            # Ignore rows that are trailing #N/A-like fillers
            str_vals = [str(v).strip().upper() if v is not None else '' for v in row]
            if all(v in ('', '#N/A', 'N#A', 'NA', 'N/A') for v in str_vals):
                row_idx += 1
                continue

            try:
                record = parse_row(row, canonical_headers)
                # Include ALL rows that have a Tag or Functional Location
                if record.get('Tag') or record.get('Functional Location'):
                    records.append(record)
            except Exception as e:
                logger.warning(f"Error parsing Data Base row {row_idx}: {str(e)}")
            finally:
                row_idx += 1
                rows_processed += 1
                # Log progress every 200 rows (reduced logging)
                if rows_processed % 200 == 0:
                    logger.info(f"Processed {rows_processed} rows, found {len(records)} items...")

        wb.close()
        logger.info(f"Successfully parsed {len(records)} items from 'Data Base' (B5:X)")
        return records
        
    except FileNotFoundError:
        raise ValueError(f"File not found: {file_path}")
    except Exception as e:
        logger.error(f"Error reading Excel file: {str(e)}", exc_info=True)
        raise ValueError(f"Failed to parse Excel file: {str(e)}")


def map_headers(headers: List[str]) -> List[str]:
    """
    Map Excel headers to canonical column names
    
    Args:
        headers: List of header strings from Excel
        
    Returns:
        List of canonical column names
    """
    canonical = []
    
    for h in headers:
        h_low = h.lower().strip()
        
        # Map known patterns
        if 'item class' in h_low:
            canonical.append('Item Class')
        elif 'description' in h_low and 'tag' not in h_low:
            canonical.append('Description')
        elif 'functional location' in h_low or 'funtional location' in h_low or ('func' in h_low and 'loc' in h_low):
            canonical.append('Functional Location')
        elif h_low == 'tag' or (h_low.startswith('tag') and 'id' not in h_low):
            canonical.append('Tag')
        elif 'last insp' in h_low:
            canonical.append('Last Insp/')
        elif 'freq' in h_low and 'sap' in h_low:
            canonical.append('Freq/ (SAP)')
        elif 'next insp' in h_low:
            canonical.append('Next Insp/')
        elif h_low == 'year':
            canonical.append('Year')
        elif 'due date' in h_low:
            canonical.append('Due Date')
        elif 'compl' in h_low and 'date' in h_low:
            canonical.append('Compl/ date')
        elif 'pmonth' in h_low and 'insp' in h_low:
            canonical.append('PMonth Insp')
        elif 'cmonth' in h_low and 'insp' in h_low:
            canonical.append('CMonth Insp')
        elif 'sece' in h_low and 'status' in h_low:
            canonical.append('SECE STATUS')
        elif 'order status' in h_low:
            canonical.append('Order Status')
        elif h_low == 'order' or ('order' in h_low and 'status' not in h_low):
            canonical.append('Order')
        elif 'delay' in h_low:
            canonical.append('Delay')
        elif 'm. item' in h_low or 'm.item' in h_low:
            canonical.append('M. Item')
        elif 'm. plan' in h_low or 'm.plan' in h_low:
            canonical.append('M. Plan')
        elif 'job done' in h_low or 'job' in h_low and 'done' in h_low:
            canonical.append('Job Done')
        elif 'days' in h_low and 'backlog' in h_low:
            canonical.append('Days in Backlog')
        elif h_low.startswith('backlog') and 'days' not in h_low:
            # Catch "Backlog?", "Backlog", "Backlog ?", "Backlog? | Days", etc.
            # but NOT "Days in Backlog"
            canonical.append('Backlog?')
        else:
            canonical.append(h)
    
    return canonical


def parse_row(row: tuple, headers: List[str]) -> Dict[str, Any]:
    """
    Parse a single row of data with proper date formatting

    Args:
        row: Tuple of cell values
        headers: List of canonical column names

    Returns:
        Dictionary with parsed data
    """
    record = {}

    for i, cell in enumerate(row):
        key = headers[i] if i < len(headers) else f"Col_{i}"

        # Due Date / Compl/ date: Excel often stores as serial number (int/float)
        if key in ('Due Date', 'Compl/ date') and cell is not None and isinstance(cell, (int, float)):
            date_str = _excel_serial_to_date_str(cell)
            if date_str:
                cell = date_str
        # Convert datetime objects to YYYY-MM-DD format
        elif isinstance(cell, datetime):
            cell = cell.strftime('%Y-%m-%d')
        # Handle date objects (from Excel date cells)
        elif hasattr(cell, 'strftime'):
            try:
                cell = cell.strftime('%Y-%m-%d')
            except Exception:
                cell = str(cell)

        # Store as string or empty
        record[key] = str(cell).strip() if cell is not None else ""
    
    # Normalize Backlog? — support multiple column names and Excel boolean/numeric
    backlog_raw = record.get('Backlog?', '') or record.get('Backlog', '')
    if isinstance(backlog_raw, bool):
        backlog_val = 'yes' if backlog_raw else 'no'
    elif isinstance(backlog_raw, (int, float)):
        backlog_val = 'yes' if backlog_raw else 'no'
    else:
        backlog_val = str(backlog_raw).strip().lower()
    record['Backlog?'] = 'Yes' if backlog_val in ('yes', 'y', 'true', '1') else 'No'

    # Parse Days in Backlog
    try:
        days_str = record.get('Days in Backlog', '0')
        days = int(float(days_str) if days_str else 0)
    except (ValueError, TypeError):
        days = 0
    record['Days in Backlog'] = days
    
    # Parse SECE status — Excel column contains "SECE" or "SCE" for critical items, blank for non-SCE
    sece_status = record.get('SECE STATUS', '').upper().strip()
    record['SECE'] = sece_status in ('SCE', 'SECE', 'YES', 'Y', 'TRUE')
    
    # Extract System and Location from Functional Location
    fl = record.get('Functional Location', '')
    parts = fl.split('/') if fl else []
    
    # Expected format: GIR/LOCATION/SYSTEM/...
    record['System'] = parts[2] if len(parts) > 2 else 'Unknown'
    record['Location'] = parts[1] if len(parts) > 1 else 'Unknown'
    # Preserve original Functional Location for UI swap
    record['Functional Location'] = fl
    
    return record


async def process_with_ai(items: List[dict]) -> List[dict]:
    """
    Process backlog items using AI (CrewAI + LangChain)
    
    Args:
        items: List of backlog items
        
    Returns:
        List of enriched dashboard items
    """
    if not AI_AVAILABLE:
        logger.warning("AI libraries not available, falling back to rule-based processing")
        return process_backlog_items(items)
    
    try:
        # Initialize LLM
        llm = ChatOpenAI(
            model="minimax/mm-m2",
            base_url="https://openrouter.ai/api/v1",
            api_key=os.getenv("OPENROUTER_API_KEY"),
            temperature=0.2,
            max_tokens=3000
        )
        
        # Create AI agent
        formatter = Agent(
            role="Oil & Gas Inspection Dashboard Specialist",
            goal="Convert raw backlog items into actionable inspection dashboard data",
            backstory=(
                "You are an expert in offshore topside integrity management with 15+ years "
                "of experience. You understand SECE classifications, risk assessment, and "
                "the criticality of inspection schedules. You create clear, concise summaries "
                "that help operations teams prioritize their work."
            ),
            llm=llm,
            verbose=False
        )
        
        # Create task
        task = Task(
            description=f"""
Analyze these {len(items)} backlog inspection items and create a dashboard-ready format.

INPUT DATA:
{json.dumps(items[:10], indent=2)}
{"... (more items truncated)" if len(items) > 10 else ""}

REQUIREMENTS:
For each item, create a JSON object with these EXACT fields:
- "Tag ID": The equipment tag (from 'Tag' field)
- "Description": Combine 'Item Class' and 'Description', max 80 characters
- "System": From parsed 'System' field
- "Location": From parsed 'Location' field  
- "Due Date": In YYYY-MM-DD format
- "Days Overdue": Integer from 'Days in Backlog'
- "SECE": "Yes" if SECE=true, else "No"
- "Status": "Overdue" if Days Overdue > 0, else "Due Soon"
- "Action": 
  * "Escalate" if Days Overdue > 90 OR (SECE and Days Overdue > 60)
  * "Schedule" otherwise
- "Risk Level":
  * "High" if Days Overdue > 90 OR (SECE and Days Overdue > 60)
  * "Medium" if Days Overdue > 30
  * "Low" otherwise
- "color": 
  * "#ffebee" (light red) for High risk
  * "#fff9c4" (light yellow) for Medium risk
  * "#e8f5e9" (light green) for Low risk

OUTPUT FORMAT:
Return ONLY valid JSON with structure: {{"table": [...]}}
No markdown, no explanations, just the JSON object.

Sort by Risk Level (High first) then Days Overdue (highest first).
""",
            expected_output="JSON object with 'table' key containing array of dashboard items",
            agent=formatter
        )
        
        # Execute crew
        crew = Crew(
            agents=[formatter],
            tasks=[task],
            verbose=False
        )
        
        logger.info("Executing AI processing...")
        result = await crew.kickoff_async()

        # CrewAI 1.x: result is CrewOutput — use .raw for string content
        result_str = (result.raw if hasattr(result, 'raw') else str(result)).strip()

        # Try to extract JSON if wrapped in markdown
        if '```json' in result_str:
            result_str = result_str.split('```json')[1].split('```')[0].strip()
        elif '```' in result_str:
            result_str = result_str.split('```')[1].split('```')[0].strip()
        
        output = json.loads(result_str)
        dashboard_items = output.get('table', [])
        
        logger.info(f"AI processing complete: {len(dashboard_items)} items")
        return dashboard_items
        
    except Exception as e:
        logger.error(f"AI processing failed: {str(e)}", exc_info=True)
        # Fallback to rule-based processing
        logger.info("Falling back to rule-based processing")
        return process_backlog_items(items)


def process_backlog_items(items: List[dict]) -> List[dict]:
    """
    Process items where Backlog='Yes'

    Args:
        items: List of all items

    Returns:
        List of backlog dashboard items
    """
    # Filter for backlog items only
    backlog_items = [item for item in items if item.get('Backlog?') == 'Yes']
    logger.info(f"Processing {len(backlog_items)} backlog items out of {len(items)} total")

    dashboard_items = []

    for item in backlog_items:
        days_overdue = item.get('Days in Backlog', 0)
        # Re-derive SECE from raw SECE STATUS to handle cached data with stale boolean
        sece_raw = str(item.get('SECE STATUS', '')).upper().strip()
        is_sece = item.get('SECE', False) or sece_raw in ('SCE', 'SECE')
        is_backlog = str(item.get('Backlog?', 'No')).lower() in ('yes', 'y', 'true', '1') or item.get('Backlog?') == 'Yes'
        
        # Determine risk level (focus on backlog items)
        if is_backlog and (days_overdue > 90 or (is_sece and days_overdue > 60)):
            risk_level = "High"
            color = "#ffebee"  # Light red
            action = "Escalate"
        elif is_backlog and days_overdue > 30:
            risk_level = "Medium"
            color = "#fff9c4"  # Light yellow
            action = "Schedule"
        else:
            risk_level = "Low"
            color = "#e8f5e9" if is_backlog else "#ffffff"  # Light green for backlog, white for non-backlog
            action = "Schedule" if is_backlog else "Monitor"
        
        # Create description
        desc_parts = []
        if item.get('Item Class'):
            desc_parts.append(item['Item Class'])
        if item.get('Description'):
            desc_parts.append(item['Description'])
        description = " - ".join(desc_parts)[:80]
        
        dashboard_items.append({
            "Tag ID": item.get('Tag', 'N/A'),
            "Category": item.get('Item Class', 'Uncategorized'),
            "Description": description,
            "System": item.get('System', 'Unknown'),
            "Location": item.get('Location', 'Unknown'),
            "Functional Location": item.get('Functional Location', 'N/A'),
            "Due Date": item.get('Due Date', 'N/A'),
            "Days Overdue": days_overdue,  # This now equals Days in Backlog
            "SECE": "Yes" if is_sece else "No",
            "SECE Status": item.get('SECE STATUS', ''),
            "Backlog?": "Yes" if is_backlog else "No",
            "Order Status": item.get('Order Status', ''),
            "Job Done": item.get('Job Done', ''),
            "Status": ("Overdue" if (is_backlog and days_overdue > 0) else ("Due Soon" if is_backlog else "Planned")),
            "Action": action,
            "Risk Level": risk_level,
            "color": color
        })
    
    # Sort by risk level and days overdue
    risk_order = {"High": 0, "Medium": 1, "Low": 2}
    dashboard_items.sort(key=lambda x: (risk_order.get(x["Risk Level"], 3), -x["Days Overdue"]))

    return dashboard_items


def _is_pending_item(item: dict, today: datetime.date) -> bool:
    """
    Shared function to determine if an item is pending.
    Pending items: Order Status in [APPR, INIT, WREA, WREL] AND (Due Date + 28 days) > Today
    
    Args:
        item: Item dictionary
        today: Today's date
        
    Returns:
        True if item is pending, False otherwise
    """
    def parse_due_date(due_date_str: str):
        """Parse Due Date string to datetime.date object."""
        if not due_date_str or due_date_str == 'N/A':
            return None
        if isinstance(due_date_str, datetime):
            return due_date_str.date()
        if isinstance(due_date_str, type(today)):
            return due_date_str
        try:
            return datetime.strptime(str(due_date_str).strip(), '%Y-%m-%d').date()
        except (ValueError, TypeError):
            try:
                return datetime.strptime(str(due_date_str).strip(), '%d/%m/%Y').date()
            except (ValueError, TypeError):
                try:
                    return datetime.strptime(str(due_date_str).strip(), '%m/%d/%Y').date()
                except (ValueError, TypeError):
                    return None
    
    order_status = item.get('Order Status', '').upper().strip()
    valid_statuses = ['APPR', 'INIT', 'WREA', 'WREL']
    
    if order_status not in valid_statuses:
        return False
    
    due_date = parse_due_date(item.get('Due Date', ''))
    if due_date is None:
        return False
    
    due_date_plus_28 = due_date + timedelta(days=28)
    return due_date_plus_28 > today


def process_performance_items(items: List[dict]) -> List[dict]:
    """
    Process items for Performance view
    Filter: Order Status = QCAP or EXDO only
    Include ALL QCAP/EXDO items regardless of other statuses

    Args:
        items: List of all items

    Returns:
        List of performance dashboard items
    """
    # Filter for performance items: Order Status = EXDO or QCAP only
    # Include ALL QCAP/EXDO items - no exclusions
    performance_items = []
    
    for item in items:
        order_status = item.get('Order Status', '').upper().strip()
        
        # Include ALL QCAP/EXDO items
        if order_status in ['QCAP', 'EXDO']:
            performance_items.append(item)
    
    logger.info(f"Processing {len(performance_items)} performance items (Order Status = QCAP/EXDO) out of {len(items)} total")

    dashboard_items = []

    for item in performance_items:
        days_in_backlog = item.get('Days in Backlog', 0)
        # Re-derive SECE from raw SECE STATUS to handle cached data with stale boolean
        sece_raw = str(item.get('SECE STATUS', '')).upper().strip()
        is_sece = item.get('SECE', False) or sece_raw in ('SCE', 'SECE')
        order_status = item.get('Order Status', 'N/A').upper().strip()
        job_done = item.get('Job Done', 'N/A')
        is_completed = 'compl' in job_done.lower()

        # All items in performance view are EXDO or QCAP (pending items excluded)
        # Determine color based on SECE status and completion
        if is_sece:
            color = "#fff9c4"  # Light yellow for SECE items
            risk_level = "Medium"
        else:
            color = "#e8f5e9"  # Light green for non-SECE items
            risk_level = "Low"
        status = "Completed" if is_completed else "In Progress"

        # Create description
        desc_parts = []
        if item.get('Item Class'):
            desc_parts.append(item['Item Class'])
        if item.get('Description'):
            desc_parts.append(item['Description'])
        description = " - ".join(desc_parts)[:80]

        dashboard_items.append({
            "Tag ID": item.get('Tag', 'N/A'),
            "Category": item.get('Item Class', 'Uncategorized'),
            "Description": description,
            "System": item.get('System', 'Unknown'),
            "Location": item.get('Location', 'Unknown'),
            "Functional Location": item.get('Functional Location', 'N/A'),
            "Due Date": item.get('Due Date', 'N/A'),
            "Completion Date": item.get('Compl/ date', 'N/A'),
            "Days in Backlog": days_in_backlog,
            "SECE": "Yes" if is_sece else "No",
            "SECE Status": item.get('SECE STATUS', ''),
            "Backlog?": item.get('Backlog?', 'No'),
            "Order Status": item.get('Order Status', 'N/A'),
            "Job Done": job_done,
            "Status": status,
            "Action": "Review",
            "Risk Level": risk_level,
            "color": color
        })

    # Sort by SECE first, then by completion date (most recent first)
    dashboard_items.sort(key=lambda x: (0 if x["SECE"] == "Yes" else 1, x.get("Completion Date", "")), reverse=True)

    return dashboard_items


def process_pending_items(items: List[dict]) -> List[dict]:
    """
    Process items for Pending view
    Filter: Order Status in [APPR, INIT, WREL, SWE] AND Backlog?="No"
    
    Args:
        items: List of all items

    Returns:
        List of pending dashboard items
    """
    # PENDING = Items with Order Status in [APPR, INIT, WREL, SWE] AND Backlog?="No"
    pending_items = []
    
    for item in items:
        order_status = item.get('Order Status', '').upper().strip()
        backlog_val = str(item.get('Backlog?', '')).lower().strip()
        is_backlog = backlog_val in ('yes', 'y', 'true', '1')
        
        # PENDING = Order Status in [APPR, INIT, WREL, SWE] AND Backlog?="No"
        # Ensure order_status is properly stripped
        order_status_clean = order_status.strip() if order_status else ''
        if order_status_clean in ['APPR', 'INIT', 'WREL', 'SWE'] and not is_backlog:
            pending_items.append(item)
    
    backlog_count = sum(1 for item in items if str(item.get('Backlog?', '')).lower() in ('yes', 'y', 'true', '1'))
    completed_count = sum(1 for item in items if item.get('Order Status', '').upper() in ['QCAP', 'EXDO'] or 'compl' in str(item.get('Job Done', '')).lower())
    total = len(items)
    
    logger.info(f"Processing {len(pending_items)} pending items (Order Status in [APPR, INIT, WREL, SWE] AND Backlog?='No') out of {total} total")
    logger.info(f"  - Total: {total}, Backlog: {backlog_count}, Completed: {completed_count}, Pending: {len(pending_items)}")

    dashboard_items = []

    for item in pending_items:
        # Calculate Days in Backlog for pending items (may be 0 or negative if not yet due)
        days_in_backlog = item.get('Days in Backlog', 0)
        # Re-derive SECE from raw SECE STATUS to handle cached data with stale boolean
        sece_raw = str(item.get('SECE STATUS', '')).upper().strip()
        is_sece = item.get('SECE', False) or sece_raw in ('SCE', 'SECE')
        order_status = item.get('Order Status', 'N/A').upper().strip()
        due_date = item.get('Due Date', 'N/A')
        
        # Determine risk level based on SECE and days
        if is_sece:
            risk_level = "Medium"
            color = "#fff9c4"  # Light yellow
        else:
            risk_level = "Low"
            color = "#e3f2fd"  # Light blue
        
        # Use the order status as-is for display (since Pending includes all non-QCAP/EXDO statuses)
        # Common statuses get readable labels, others are shown as-is
        status_map = {
            'APPR': 'Approved',
            'WREL': 'Work Released',
            'INIT': 'Initiated',
            'SWE': 'Site Work Execution',
            'PREL': 'Pre-Release',
            'TECO': 'Technically Complete',
            'CLSD': 'Closed',
            'CNCL': 'Cancelled',
            'DLFL': 'Deferred',
            'REL': 'Released',
            'PRT': 'Partially Released',
            'CRTD': 'Created'
        }
        # Use mapped status if available, otherwise use the order status as-is
        status = status_map.get(order_status, order_status if order_status and order_status != 'N/A' else 'Pending')
        
        # Create description
        desc_parts = []
        if item.get('Item Class'):
            desc_parts.append(item['Item Class'])
        if item.get('Description'):
            desc_parts.append(item['Description'])
        description = " - ".join(desc_parts)[:80]

        dashboard_items.append({
            "Tag ID": item.get('Tag', 'N/A'),
            "Category": item.get('Item Class', 'Uncategorized'),
            "Description": description,
            "System": item.get('System', 'Unknown'),
            "Location": item.get('Location', 'Unknown'),
            "Functional Location": item.get('Functional Location', 'N/A'),
            "Due Date": due_date,
            "Days in Backlog": days_in_backlog,
            "SECE": "Yes" if is_sece else "No",
            "SECE Status": item.get('SECE STATUS', ''),
            "Backlog?": "No",
            "Order Status": item.get('Order Status', 'N/A'),
            "Job Done": item.get('Job Done', 'N/A'),
            "Status": status,
            "Action": "Monitor",
            "Risk Level": risk_level,
            "color": color
        })

    # Sort by SECE first, then by due date
    dashboard_items.sort(key=lambda x: (0 if x["SECE"] == "Yes" else 1, x.get("Due Date", "")))

    return dashboard_items
